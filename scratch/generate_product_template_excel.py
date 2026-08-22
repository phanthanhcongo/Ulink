import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_excel_template():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # SHEET 1: Product Data Template
    # ----------------------------------------------------
    ws_main = wb.active
    ws_main.title = "Product_Data_Template"
    ws_main.views.sheetView[0].showGridLines = True

    # ----------------------------------------------------
    # SHEET 2: Master Data & Instructions
    # ----------------------------------------------------
    ws_master = wb.create_sheet(title="Master_Data_Lists")
    ws_master.views.sheetView[0].showGridLines = True

    # Styles
    navy_header_fill = PatternFill(start_color="112842", end_color="112842", fill_type="solid")
    blue_accent_fill = PatternFill(start_color="1769E2", end_color="1769E2", fill_type="solid")
    light_blue_fill = PatternFill(start_color="EAF1FA", end_color="EAF1FA", fill_type="solid")
    sample_row_fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")

    font_title = Font(name="Segoe UI", size=14, bold=True, color="112842")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="536174")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_master_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10, color="172033")
    font_sample = Font(name="Segoe UI", size=10, italic=True, color="4D5969")
    font_bold_body = Font(name="Segoe UI", size=10, bold=True, color="172033")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    thin_border_side = Side(border_style="thin", color="D7E1ED")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=Side(border_style="medium", color="1769E2"))

    # Populate Sheet 2: Master Data
    ws_master["A1"] = "DANH MỤC CHA (PARENT CATEGORY)"
    ws_master["B1"] = "DANH MỤC CON (SUB-CATEGORY)"
    ws_master["C1"] = "ĐƠN VỊ TÍNH (UNITS)"
    ws_master["D1"] = "TRẠNG THÁI (STATUS)"

    for col_letter in ["A", "B", "C", "D"]:
        cell = ws_master[f"{col_letter}1"]
        cell.font = font_master_header
        cell.fill = navy_header_fill
        cell.alignment = align_center

    parent_cats = ["Phòng sạch", "Đóng gói"]
    sub_cats = [
        "Găng tay phòng sạch",
        "Khăn lau phòng sạch",
        "Trang phục phòng sạch",
        "Khẩu trang phòng sạch",
        "Thảm & Dụng cụ ESD",
        "Hóa chất phòng sạch",
        "Màng co PE",
        "Màng quấn Pallet",
        "Túi PE / ESD",
        "Băng keo công nghiệp & HVAC",
        "Bao bì định hình & Phụ trợ"
    ]
    units = ["Đôi", "Cuộn", "Hộp", "Bộ", "Thùng", "Kg", "Tấm", "Gói", "Can", "Chai", "Cái"]
    statuses = ["Published", "Draft"]

    for i, val in enumerate(parent_cats, start=2):
        ws_master[f"A{i}"] = val
    for i, val in enumerate(sub_cats, start=2):
        ws_master[f"B{i}"] = val
    for i, val in enumerate(units, start=2):
        ws_master[f"C{i}"] = val
    for i, val in enumerate(statuses, start=2):
        ws_master[f"D{i}"] = val

    # Setup Main Sheet Header
    ws_main.merge_cells("A1:S1")
    ws_main["A1"] = "ULINK INDUSTRIES — BẢNG TEMPLATE NHẬP DỮ LIỆU SẢN PHẨM (PRODUCT DATA TEMPLATE)"
    ws_main["A1"].font = font_title
    ws_main["A1"].alignment = align_left

    ws_main.merge_cells("A2:S2")
    ws_main["A2"] = "Vui lòng nhập dữ liệu theo đúng định dạng. Các cột có dấu (*) là bắt buộc. Sử dụng menu thả xuống (Dropdown) để chọn Danh mục & Đơn vị tính."
    ws_main["A2"].font = font_subtitle
    ws_main["A2"].alignment = align_left

    ws_main.row_dimensions[1].height = 28
    ws_main.row_dimensions[2].height = 20
    ws_main.row_dimensions[4].height = 36

    headers = [
        "STT",
        "Danh mục Cha (*)",
        "Danh mục Con (*)",
        "Tên sản phẩm (*)",
        "Mã SKU (*)",
        "Thương hiệu",
        "Đơn vị tính (*)",
        "Quy cách đóng gói",
        "Thuộc tính biến thể\n(Size / Màu / Quy cách)",
        "MOQ (*)",
        "Giá sỉ cơ bản\n(VNĐ)",
        "Giá nấc 1\n(500 - 999)",
        "Giá nấc 2\n(1.000 - 2.999)",
        "Giá nấc 3\n(>= 3.000)",
        "Thông số kỹ thuật chính\n(Specifications)",
        "Mô tả ngắn sản phẩm",
        "Link Ảnh\n(Image)",
        "Link File PDF\n(TDS/MSDS)",
        "Trạng thái"
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws_main.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = border_header

    # Sample rows data
    sample_data = [
        [
            1, "Phòng sạch", "Găng tay phòng sạch", "Găng tay phòng sạch Nitrile 12 inch", "ULK-GLV-NIT-12",
            "ULink / Ansell", "Đôi", "100 cái/hộp, 10 hộp/thùng", "Kích cỡ: S, M, L, XL | Màu sắc: Trắng, Xanh",
            500, 2500, 2500, 2300, 2100,
            "Cấp sạch: ISO Class 5 | Chất liệu: Nitrile 100% | Độ dài: 12 inch",
            "Găng tay Nitrile chống tĩnh điện không bột chuyên dùng cho phòng sạch sản xuất linh kiện điện tử.",
            "gang-tay-nitrile.png", "TDS-Nitrile-12inch.pdf", "Published"
        ],
        [
            2, "Phòng sạch", "Khăn lau phòng sạch", "Khăn lau phòng sạch Polyester 100% 9x9 inch", "ULK-WIP-POLY-9",
            "Texwipe", "Gói", "150 tờ/gói, 10 gói/thùng", "Kích thước: 9x9 inch | Mép cắt: Laser",
            50, 250000, 250000, 235000, 220000,
            "Chất liệu: 100% Polyester Filament | Mép cắt: Cắt Laser không xơ sợi | Cấp sạch: ISO Class 5",
            "Khăn lau phòng sạch thấm hút cao, không xơ bụi, chuyên lau bề mặt kính và linh kiện tinh vi.",
            "khan-lau-polyester.png", "TDS-Wiper-Polyester.pdf", "Published"
        ],
        [
            3, "Đóng gói", "Màng co PE", "Màng quấn Pallet PE quấn tay 3.0kg", "ULK-PKG-STRETCH-3K",
            "ULink", "Cuộn", "4 cuộn/thùng", "Trọng lượng: 3.0 kg | Độ dày: 17 micron | Độ giãn: 250%",
            100, 120000, 120000, 112000, 105000,
            "Chất liệu: Hạt nhựa LLDPE nguyên sinh | Độ dày: 17 mic | Trọng lượng lõi: 0.5 kg",
            "Màng quấn Pallet bám dính cao, chống đâm thủng, co bền chắc giúp cố định kiện hàng vận chuyển.",
            "mang-quan-pallet.png", "TDS-Stretch-Film.pdf", "Published"
        ],
        [
            4, "Đóng gói", "Túi PE / ESD", "Túi chống tĩnh điện ESD Shielding Bag", "ULK-PKG-ESD-SHD",
            "Desco", "Túi", "100 túi/xấp, 1.000 túi/thùng", "Kích thước: 15x20 cm, 20x30 cm | Kiểu: Miệng phẳng, Zipper",
            1000, 3500, 3500, 3200, 2900,
            "Điện trở bề mặt: <10^11 Ω | Cấu trúc: 4 lớp PET/AL/NY/PE | Tiêu chuẩn: ANSI/ESD S541",
            "Túi ESD Shielding bảo vệ an toàn linh kiện điện tử nhạy cảm khỏi nguy cơ phóng điện ESD.",
            "tui-esd-shielding.png", "TDS-ESD-Bag.pdf", "Published"
        ]
    ]

    # Populate Sample Rows (Rows 5 to 8)
    for r_idx, row_val in enumerate(sample_data, start=5):
        ws_main.row_dimensions[r_idx].height = 42
        for c_idx, val in enumerate(row_val, start=1):
            cell = ws_main.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_sample
            cell.fill = sample_row_fill
            cell.border = border_all
            
            # Formats & Alignments
            if c_idx in [1, 10, 19]:
                cell.alignment = align_center
            elif c_idx in [11, 12, 13, 14]:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            else:
                cell.alignment = align_left

    # Empty user entry rows (Rows 9 to 60)
    for r_idx in range(9, 61):
        ws_main.row_dimensions[r_idx].height = 24
        for c_idx in range(1, 20):
            cell = ws_main.cell(row=r_idx, column=c_idx)
            cell.font = font_body
            cell.border = border_all
            if c_idx in [1, 10, 19]:
                cell.alignment = align_center
            elif c_idx in [11, 12, 13, 14]:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            else:
                cell.alignment = align_left

    # Add Data Validation (Dropdowns) for rows 5 to 100
    dv_parent = DataValidation(type="list", formula1="Master_Data_Lists!$A$2:$A$3", allow_blank=True)
    dv_sub = DataValidation(type="list", formula1="Master_Data_Lists!$B$2:$B$12", allow_blank=True)
    dv_unit = DataValidation(type="list", formula1="Master_Data_Lists!$C$2:$C$12", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="Master_Data_Lists!$D$2:$D$3", allow_blank=True)

    ws_main.add_data_validation(dv_parent)
    ws_main.add_data_validation(dv_sub)
    ws_main.add_data_validation(dv_unit)
    ws_main.add_data_validation(dv_status)

    dv_parent.add("B5:B100")
    dv_sub.add("C5:C100")
    dv_unit.add("G5:G100")
    dv_status.add("S5:S100")

    # Set Column Widths
    col_widths = {
        "A": 8,   # STT
        "B": 18,  # Parent Cat
        "C": 26,  # Sub Cat
        "D": 38,  # Product Name
        "E": 22,  # SKU
        "F": 18,  # Brand
        "G": 14,  # Unit
        "H": 26,  # Pack size
        "I": 36,  # Attributes
        "J": 12,  # MOQ
        "K": 18,  # Base Price
        "L": 18,  # Tier 1
        "M": 18,  # Tier 2
        "N": 18,  # Tier 3
        "O": 45,  # Specs
        "P": 45,  # Short desc
        "Q": 24,  # Image link
        "R": 24,  # PDF link
        "S": 14   # Status
    }
    for col_letter, w in col_widths.items():
        ws_main.column_dimensions[col_letter].width = w

    # Master sheet column widths
    ws_master.column_dimensions["A"].width = 32
    ws_master.column_dimensions["B"].width = 36
    ws_master.column_dimensions["C"].width = 24
    ws_master.column_dimensions["D"].width = 20

    # Save to file
    out_dir = r"c:\Users\thanh\Desktop\PathtechProject\ulink-b2b-platform\figmaPreview\CustomerDocs"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "ULink_Product_Data_Template.xlsx")
    wb.save(out_path)
    print(f"Excel template successfully created at: {out_path}")

if __name__ == "__main__":
    create_excel_template()
