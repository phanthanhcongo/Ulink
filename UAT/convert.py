import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def parse_md_tables(text):
    """Parse markdown tables from text, return list of (section_title, rows, headers)"""
    lines = text.split('\n')
    sections = []
    current_section = "Misc"
    i = 0
    while i < len(lines):
        m = re.match(r'^###+\s+(.+)', lines[i])
        if m:
            current_section = m.group(1).strip()
        i += 1

    # Find all table blocks
    tables = []
    i = 0
    while i < len(lines):
        if lines[i].strip().startswith('|') and '|' in lines[i][1:]:
            header_line = lines[i]
            sep_line = lines[i+1] if i+1 < len(lines) else ''
            if re.match(r'^\|[\s\-:|]+\|$', sep_line):
                headers = [h.strip() for h in header_line.strip().strip('|').split('|')]
                rows = []
                j = i + 2
                while j < len(lines) and lines[j].strip().startswith('|'):
                    cells = [c.strip() for c in lines[j].strip().strip('|').split('|')]
                    # Pad cells to match headers
                    while len(cells) < len(headers):
                        cells.append('')
                    rows.append(cells)
                    j += 1
                # Find nearest section title above
                section = "Misc"
                for k in range(i-1, -1, -1):
                    m = re.match(r'^##+\s+(.+)', lines[k])
                    if m:
                        section = m.group(1).strip()
                        break
                # Clean section from "### 1.1 Title" to just "Title"
                section = re.sub(r'^[\d.]+\s*', '', section)
                tables.append((section, headers, rows))
                i = j
                continue
        i += 1
    return tables

def clean_cell(text):
    """Clean cell text: replace <br> with newline, strip"""
    text = re.sub(r'<br\s*/?>', '\n', text)
    text = text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    return text.strip()

# Read file
with open(r'C:\Users\thanh\Desktop\PathtechProject\ulink-b2b-platform\UAT\UAT_checklist.md', 'r', encoding='utf-8') as f:
    md_text = f.read()

# Fix encoding issues
md_text = md_text.encode('utf-8', errors='replace').decode('utf-8')

tables = parse_md_tables(md_text)

wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(vertical='top', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

sheet_map = {}
for section, headers, rows in tables:
    sheet_name = re.sub(r'[/\\*?\[\]:]', '', section)[:31]  # Excel sheet name max 31 chars
    if not sheet_name:
        sheet_name = "Sheet"
    
    # Handle duplicates
    if sheet_name in sheet_map:
        sheet_map[sheet_name] += 1
        sheet_name = f"{sheet_name[:28]}_{sheet_map[sheet_name]}"
    else:
        sheet_map[sheet_name] = 1
    
    ws = wb.create_sheet(title=sheet_name)
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=clean_cell(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Write rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, cell_text in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=clean_cell(cell_text))
            cell.border = thin_border
            if col_idx in (3, 4, 5):  # Desktop, Tablet, Mobile columns
                cell.alignment = center_align
            else:
                cell.alignment = cell_align
    
    # Auto-size columns (with cap)
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                # Use longest line for width calculation
                lines = val.split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
        # Cap at 80
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = min(max_len + 2, 80)

output_path = r'C:\Users\thanh\Desktop\PathtechProject\ulink-b2b-platform\UAT\UAT_checklist.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
print(f'Sheets: {wb.sheetnames}')
